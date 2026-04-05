import openpyxl
from config import get_ruta_excel


def leer_comidas_excel(fecha_inicio, fecha_fin) -> list[dict]:
    wb = openpyxl.load_workbook(get_ruta_excel())
    ws = wb["COMIDAS"]
    wb.close()

    from datetime import datetime
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha_celda, desayuno, almuerzo, cena, extras = row
        if fecha_celda is None:
            continue
        if isinstance(fecha_celda, datetime):
            fecha_row = fecha_celda.date()
        else:
            fecha_row = fecha_celda
        if fecha_row < fecha_inicio or fecha_row > fecha_fin:
            continue

        def parsear(valor):
            if valor is None or str(valor).strip() == "":
                return []
            return [a.strip() for a in str(valor).split("/") if a.strip()]

        datos.append({
            "fecha": fecha_row,
            "desayuno": parsear(desayuno),
            "almuerzo": parsear(almuerzo),
            "cena": parsear(cena),
            "extras": parsear(extras),
        })

    datos.sort(key=lambda x: x["fecha"])
    return datos


def leer_dia_excel(fecha) -> dict:
    wb = openpyxl.load_workbook(get_ruta_excel())
    ws = wb["COMIDAS"]
    wb.close()

    from datetime import datetime
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha_celda, desayuno, almuerzo, cena, extras = row
        if fecha_celda is None:
            continue
        if isinstance(fecha_celda, datetime):
            fecha_row = fecha_celda.date()
        else:
            fecha_row = fecha_celda
        if fecha_row == fecha:
            def parsear(valor):
                if valor is None or str(valor).strip() == "":
                    return []
                return [a.strip() for a in str(valor).split("/") if a.strip()]
            return {
                "desayuno": parsear(desayuno),
                "almuerzo": parsear(almuerzo),
                "cena": parsear(cena),
                "extras": parsear(extras),
            }
    return {"desayuno": [], "almuerzo": [], "cena": [], "extras": []}


def leer_alimentos_referencia() -> dict:
    wb = openpyxl.load_workbook(get_ruta_excel())
    ws = wb["ALIMENTOS"]
    wb.close()

    referencia = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre, tipo, proteina, grasa, carbos, kcal = row
        if nombre is None:
            continue
        referencia[str(nombre).strip()] = {
            "tipo": str(tipo).strip() if tipo else None,
            "proteina": float(proteina) if proteina is not None else None,
            "grasa": float(grasa) if grasa is not None else None,
            "carbos": float(carbos) if carbos is not None else None,
            "kcal": float(kcal) if kcal is not None else None,
        }
    return referencia


def guardar_comidas_excel(fecha, desayuno, almuerzo, cena, extras) -> bool:
    from datetime import datetime
    wb = openpyxl.load_workbook(get_ruta_excel())
    ws = wb["COMIDAS"]

    fila_existente = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        fecha_celda = row[0]
        if fecha_celda is None:
            continue
        if isinstance(fecha_celda, datetime):
            fecha_celda = fecha_celda.date()
        if fecha_celda == fecha:
            fila_existente = row_idx
            break

    if fila_existente:
        if desayuno is not None:
            ws.cell(row=fila_existente, column=2).value = desayuno
        if almuerzo is not None:
            ws.cell(row=fila_existente, column=3).value = almuerzo
        if cena is not None:
            ws.cell(row=fila_existente, column=4).value = cena
        if extras is not None:
            ws.cell(row=fila_existente, column=5).value = extras
    else:
        ws.append([fecha, desayuno, almuerzo, cena, extras])

    wb.save(get_ruta_excel())
    wb.close()
    return True
